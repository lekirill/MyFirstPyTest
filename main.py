import openpyxl as xl_tool
wb = xl_tool.load_workbook('./Слова_англ.xlsx')
print(wb.sheetnames)
ws = wb['Лист1']
print(ws['A3'].value)
print(ws['A4'].value)
wb.close()